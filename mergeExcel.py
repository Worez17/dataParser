#-*- coding: utf8 -*-
import xlrd
import openpyxl

fnames = ["G:\data_source\\194434611.xlsx", "G:\data_source\\200932002.xlsx", "G:\data_source\\203328886.xlsx", "G:\data_source\\205834477.xlsx"
          , "G:\data_source\\212437013.xlsx", "G:\data_source\\214712057.xlsx", "G:\data_source\\215024555.xlsx", "G:\data_source\\215357148.xlsx"
          , "G:\data_source\\215817158.xlsx", "G:\data_source\\220049305.xlsx"
         ]
row_merge = 0

outwb = openpyxl.Workbook()
outws = outwb.create_sheet('sheet')

for fname in fnames:
    data = xlrd.open_workbook(fname).sheets()[0]
    row0 =data.row_values(0)
    k = 0
    # 表头信息
    if row_merge == 0:
        outws.cell(row=1, column=1).value = row0[0]
        outws.cell(row=1, column=2).value = row0[1]
        outws.cell(row=1, column=3).value = row0[2]
        outws.cell(row=1, column=4).value = row0[3]
        outws.cell(row=1, column=5).value = row0[5]
        outws.cell(row=1, column=6).value = row0[6]
        outws.cell(row=1, column=7).value = row0[7]
        k = 1
    # 写入表数据
    for i in range(1, data.nrows):
        row = data.row_values(i)
        outws.cell(row=row_merge + i + k, column=1).value = row[0]
        outws.cell(row=row_merge + i + k, column=2).value = row[1]
        outws.cell(row=row_merge + i + k, column=3).value = row[2]
        outws.cell(row=row_merge + i + k, column=4).value = row[3]
        outws.cell(row=row_merge + i + k, column=5).value = row[5]
        outws.cell(row=row_merge + i + k, column=6).value = row[6]
        outws.cell(row=row_merge + i + k, column=7).value = row[7]
    row_merge = row_merge + data.nrows
    print fname

outwb.save('G:result\\allGuest.xlsx')
print 'OK'
